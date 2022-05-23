from flask import *
from moviepy.editor import *
import os
import openpyxl
import yt_dlp
# from __future__ import unicode_literals



app = Flask(__name__,template_folder='template',static_folder='static')


INPUT_FOLDER = 'Inputs'
OUTPUT_FOLDER = 'static/Outputs'
DOWNLOAD_FOLDER = 'Downloaded'
PARENT_DIR = os.getcwd()
input_path = os.path.join(PARENT_DIR,INPUT_FOLDER)
output_path = os.path.join(PARENT_DIR,OUTPUT_FOLDER)
downloaded_path = os.path.join(PARENT_DIR,DOWNLOAD_FOLDER)



'''
Downloades the Youtube Videos on specified DOWNLOAD_FOLDER
Returns list of Titles(With IDS) 
'''
def download_video(urls):
    # os.chdir(downloaded_path)
    titles = []
    i=1
    for url in urls:
        if url == None:
            continue

        ydl_opts = {
            'format': 'mp4',
            'outtmpl': f'{downloaded_path}/{i}_%(title)s.%(ext)s'
        }

        link_of_the_video = str(url)
        zxt = link_of_the_video.strip()

        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(zxt, download=True)
            video_title = info.get('title', None)
            ydl.prepare_filename(info)
            titles.append(f'{i}_{video_title}.mp4')

        i+=1

    # os.chdir(PARENT_DIR)
    return titles



def edit_video(titles,start,end):

    print("Titles in edit_video",titles)

    os.chdir(downloaded_path)
    video_objects = []
    for title in titles:
        vid = VideoFileClip(title)
        video_objects.append(vid)

    # os.chdir(output_path)

    for i in range(len(video_objects)):
        
        clip = video_objects[i].subclip(start[i],end[i])
        clip.write_videofile(f'{output_path}/{titles[i]}')

    # os.chdir(PARENT_DIR)
    
    return titles




# Landing Page
@app.route('/', methods=['GET', 'POST'])
def main():
    if request.method=="POST":

        f = request.files['file']  
        f.save(os.path.join(input_path,f.filename)) 

        URL_COLUMN_NAME = str(request.form['URL_COLUMN_NAME'])
        START_TIME =str( request.form['START_TIME'])
        END_TIME = str(request.form['END_TIME'])
        
        os.chdir(input_path)
        wb_obj = openpyxl.load_workbook(str(f.filename))
        sheet = wb_obj.active

        urls = []
        start = []
        end = []

        for column_cell in sheet.iter_cols(1, sheet.max_column):  
            if column_cell[0].value == URL_COLUMN_NAME:    
                for data in column_cell[1:]:   
                    urls.append(data.value)
    
            elif column_cell[0].value == START_TIME:    
                for data in column_cell[1:]:   
                    start.append(data.value)
              
            elif column_cell[0].value == END_TIME:    
                for data in column_cell[1:]:   
                    end.append(data.value)
               

        print(URL_COLUMN_NAME,START_TIME,END_TIME)
        print(urls,start,end)

        titles = download_video(urls)
        print("Videos Downloaded ")

        titles = edit_video(titles,start,end)

        # dir = downloaded_path
        # for f in os.listdir(dir):
        #     os.remove(os.path.join(dir, f))
        # print("Videos Subclipped.")



        return render_template('Outputindex.html', titles=titles, output_folder = OUTPUT_FOLDER)


    
    return render_template('index.html')


if __name__ == "__main__":
    app.secret_key = "xyz" 
    app.run(debug=True)